from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import pytesseract
import subprocess
import pathlib
import tempfile
import os

SOFFICE = os.getenv("SOFFICE_PATH")
if not SOFFICE:
    if os.name == 'nt':  # Windows
        SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"
    else:  # Linux / Docker
        SOFFICE = "/usr/bin/soffice"


def extract_doc(archivo):
    """
    Convierte un archivo .doc a texto mediante LibreOffice en modo headless.

    SEGURIDAD [A2-4.1]: Se reemplazó el directorio compartido (outdir = archivo.parent)
    por un directorio temporal exclusivo por ejecución (tempfile.TemporaryDirectory).
    Esto elimina la condición de carrera donde múltiples conversiones concurrentes
    en el mismo directorio podían cruzar contenidos al buscar el .txt "más reciente".
    Cada conversión ahora opera en su propio directorio aislado y determinístico.
    """
    archivo = pathlib.Path(archivo)

    with tempfile.TemporaryDirectory() as tmp_dir:
        subprocess.run([
            SOFFICE,
            "--headless",
            "--convert-to", "txt:Text",
            str(archivo),
            "--outdir", tmp_dir
        ], check=True, timeout=60)

        # Nombre de salida determinístico: mismo stem que el archivo fuente
        expected_txt = pathlib.Path(tmp_dir) / (archivo.stem + ".txt")

        if not expected_txt.exists():
            # Fallback: cualquier .txt en el directorio temporal (solo este proceso)
            txt_files = list(pathlib.Path(tmp_dir).glob("*.txt"))
            if not txt_files:
                raise FileNotFoundError("LibreOffice no generó ningún archivo .txt")
            expected_txt = txt_files[0]

        return expected_txt.read_text(encoding="utf-8", errors="ignore")


def extract_text_from_file(file_path, extension):
    try:
        if extension == "pdf":
            reader = PdfReader(file_path)
            return "\n".join(page.extract_text() or "" for page in reader.pages)

        elif extension == "docx":
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs)

        elif extension == "doc":
            return extract_doc(file_path)

        elif extension in ("xls", "xlsx"):
            wb = load_workbook(file_path, data_only=True)
            text = []
            for sheet in wb:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            text.append(str(cell.value))
            return "\n".join(text)

        elif extension in ("png", "jpg", "jpeg"):
            img = Image.open(file_path)
            return pytesseract.image_to_string(img)

        else:
            return ""

    except Exception:
        return ""
